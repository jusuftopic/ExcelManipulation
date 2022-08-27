import openpyxl as xl
from openpyxl.chart import BarChart, Reference
from utils.enums import Types, Languages


def do_excel_manipulation():
    try:
        programmers_file = xl.load_workbook('programmers.xlsx')
        programmers_table = programmers_file.active
        # check if there is programmers table
        if is_table_active(programmers_table):
            remove_invalid_names(programmers_table)
            correct_type(programmers_table)
            filter_relevant_languages(programmers_table)
            draw_barchart(programmers_table)

        programmers_file.save("valid_programmers_table.xlsx")

    except FileNotFoundError:
        print("File not found. Please check the file name")


def is_table_active(table):
    """
    Check if table in excel file exists

    :param table: Excel table
    :return: Boolean value whether table exists or not
    """
    if table is None:
        print("Table does not exists in excel file")
        return

    return True


def remove_invalid_names(table):
    """
    Remove all invalid programmer names

    :param table: Excel table
    """
    for row in range(2, table.max_row):
        firstname = table.cell(row, 1).value
        lastname = table.cell(row, 2).value

        if firstname == "-" or lastname == "-":
            table.delete_rows(row)


def correct_type(table):
    """
    Correct all lower-case type values and replace it with upper-case values

    :param table: Excel table
    """
    types = [t.value for t in Types]
    for row in range(2, table.max_row):
        type_cell = table.cell(row, 3)
        upper_case_value = type_cell.value
        if upper_case_value is not None and upper_case_value.upper() in types:
            type_cell.value = upper_case_value
        else:
            table.delete_rows(row)


def filter_relevant_languages(table):
    """
    Filter only contributions in chart relevant languages

    :param table: Excel table
    """
    languages = [language.value for language in Languages]
    for row in range(2, table.max_row):
        language_cell = table.cell(row, 4).value
        if language_cell is not  None and language_cell in languages:
            table.delete_rows(row)



def draw_barchart(table):
    """
    Draws bar chart based on languages and LOG

    :param table: Excel table
    """
    barchart = BarChart()
    barchart.type = "col"
    barchart.style = 10
    barchart.title = "Languages chart"

    data = Reference(table, 5, 2, 5, table.max_row)
    languages = Reference(table, 4, 2, 4, table.max_row)

    barchart.add_data(data, titles_from_data=True)
    barchart.set_categories(languages)
    barchart.shape = 4

    table.add_chart(barchart)


do_excel_manipulation()